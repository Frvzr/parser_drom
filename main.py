# -*- coding: utf-8 -*-

from config import urls
import requests
from bs4 import BeautifulSoup as bs
import re
import openpyxl
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import time
import datetime

pages = []


def get_request(url):
    headers = {"Accept": "*/*",
               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"}
    page = requests.get(url, headers=headers)
    return parse_pages(page)


def parse_pages(page):
    global pages
    soup = bs(page.text, 'html.parser')
    pages_ = soup.find_all('a', class_='css-1jjais5 ena3a8q0')
    for page in pages_:
        page_ = page.get('href')
        if page_ not in pages:
            pages.append(page_)
    if len(pages) < 100:
        lst = pages_[-1].get('href')
        if pages[-1] == lst:
            return get_request(pages[-1])
    print(len(pages))
    return soup_data(pages)


def soup_data(pages):
    cars = []
    headers = {"Accept": "*/*",
               "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/103.0.0.0 Safari/537.36"}
    for url in pages:
        page = requests.get(url, headers=headers)
        soup = bs(page.text, 'html.parser')
        cars.append(soup.find_all('a', class_='css-ck6dgx ewrty961'))
    cars_ = sum(cars, [])
    print(len(cars_))
    return collect_data(cars_)


def collect_data(cars):
    cars_dict = {}
    fuel_list = ['бензин', 'дизель', 'электро', 'гибрид', 'гбо']
    transmissions_list = ['автомат', 'акпп', 'робот', 'вариатор', 'механика']
    drive_list = ['4wd', 'передний', 'задний']
    fuel = ' '
    transmission = ' '
    drive = ' '
    for car in cars:
        description = []
        link = car.get('href')
        id = re.findall('\d{8}', str(link))
        car_name = car.find(
            'span', {'data-ftid': 'bull_title'}).text.split(',')

        price = car.find('span', {'data-ftid': 'bull_price'}).text
        price_ = re.sub(r"\s+", "", price)

        desc = car.find_all('span', class_='css-1l9tp44 e162wx9x0')
        for i in desc:
            description.append(i.text.strip(','))

        try:
            engine = re.findall(r"\d{1}\.\d{1}", str(description))
            engine = ''.join(map(str, engine))
        except:
            engine = ' '

        try:
            hp = re.findall(r"\d{2,3} [л]\.[с]", str(description))
            hp = ''.join(map(str, hp))
        except:
            hp = ' '

        for fuel_type in description:
            if fuel_type.lower() in fuel_list:
                fuel = fuel_type
                break

        for transmissions_type in description:
            if transmissions_type.lower() in transmissions_list:
                transmission = transmissions_type
                break

        for drive_type in description:
            if drive_type.lower() in drive_list:
                drive = drive_type
                break

        try:
            mileage = re.findall(r"[0-9]+ тыс\. км", str(description))
            mileage = ''.join(map(str, mileage))
        except:
            mileage = ' '

        try:
            city = car.find('span', {'data-ftid': 'bull_location'}).text
        except:
            city = ' '

        cars_dict.setdefault(
            *id, [*car_name, engine, hp, fuel, transmission, drive, mileage, link, city, price_])
    print(len(cars_dict))
    return create_file(cars_dict)


def create_file(cars_dict):
    id_list = []
    title_list = []
    today = f"{datetime.datetime.today().strftime('%Y-%m-%d')}"

    try:
        wb = openpyxl.load_workbook(f'C:\\Users\\user\\Desktop\\cars.xlsx')
        ws = wb.active
    except:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['id', 'car_name', 'year', 'engine', 'hp', 'fuel',
                  'transmissions', 'drive', 'mileage', 'link', 'city'])

    maxi_column = ws.max_column
    maxi_row = ws.max_row

    for row_cells in ws.iter_rows(min_row=1, max_row=1):
        for row_cell in row_cells:
            title_list.append(row_cell.value)
    if today not in title_list:
        ws.cell(row=1, column=maxi_column+1).value = today
        title_list.append(today)
        print(title_list)

    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:
            id_list.append(cell.value)
    print(len(id_list))

    for row, (key, values) in enumerate(cars_dict.items(), start=maxi_row+1):
        row = ws.max_row+1
        if str(key) in id_list and today in title_list:
            idx = id_list.index(key)
            ws.cell(row=idx+1, column=title_list.index(today)+1
                    ).value = values[10]
        else:
            ws.cell(row=row, column=title_list.index(
                today)+1).value = values[10]
            ws[f'A{row}'] = key
            ws[f'B{row}'] = values[0]
            ws[f'C{row}'] = values[1]
            ws[f'D{row}'] = values[2]
            ws[f'E{row}'] = values[3]
            ws[f'F{row}'] = values[4]
            ws[f'G{row}'] = values[5]
            ws[f'H{row}'] = values[6]
            ws[f'I{row}'] = values[7]
            ws[f'J{row}'] = values[8]
            ws[f'K{row}'] = values[9]
    return format_file(wb, ws)


def format_file(wb, ws):
    tab = Table(displayName="Table1", ref=f"A1:" +
                get_column_letter(ws.max_column) + str(ws.max_row))
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    try:
        ws.add_table(tab)
        col_dimensions = {'A': 15, 'B': 20, 'C': 12, 'D': 12, 'E': 15,
                          'F': 15, 'G': 15, 'H': 15, 'I': 15, 'J': 75, 'K': 20, 'L': 20}
        for key, value in col_dimensions.items():
            ws.column_dimensions[key].width = value
    except:
        del ws.tables["Table1"]
        ws.add_table(tab)
        last_column = get_column_letter(ws.max_column)
        ws.column_dimensions[last_column].width = 20

    for i in range(ws.max_row + 1):
        ws.row_dimensions[i].height = 20
    return save_excel(wb)


def save_excel(wb):
    try:
        f_name = f'C:\\Users\\user\\Desktop\\cars.xlsx'
        wb.save(f_name)
        message = f'данные записаны в файл'
    except:
        message = f'ошибка записи в файл'
    finally:
        wb.close
        print(message)


def main(url):
    start = time.time()
    get_request(url)
    end = time.time()
    print(end - start)


if __name__ == '__main__':
    for url in urls:
        main(url)
